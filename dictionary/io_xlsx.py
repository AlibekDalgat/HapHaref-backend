"""XLSX export/import of dictionary entries.

The format is round-trippable: ``export_words_xlsx`` produces a workbook that
``import_words_xlsx`` can read back. One row per word. Nested data lives in two
cells:

  * ``translations`` — строки вида ``lang|text`` (по одной на строку в ячейке);
  * ``examples``     — строки вида ``arabic|cyrillic|translation|source``.

Import is an **upsert by ``headword_arabic``** (повторный импорт обновляет, а не
плодит дубли). Columns are matched by header name, so a parsed dictionary may
provide a subset/reordering — only ``headword_arabic`` is required.
"""
from io import BytesIO

from openpyxl import Workbook, load_workbook

from .models import PartOfSpeech, Root, Translation, UsageExample, Word

COLUMNS = [
    "headword_arabic",
    "transliteration_latin",
    "transliteration_cyrillic",
    "pronunciation",
    "part_of_speech",
    "root_morpheme",
    "root_translit",
    "root_origin",
    "definition",
    "etymology",
    "notes",
    "status",
    "translations",
    "examples",
]

_POS_VALUES = {c[0] for c in PartOfSpeech.choices}
_STATUS_VALUES = {c[0] for c in Word.Status.choices}
_LANG_VALUES = {c[0] for c in Translation.Language.choices}
_ORIGIN_VALUES = {c[0] for c in Root.Origin.choices}


# --- export -----------------------------------------------------------------

def _join_translations(word) -> str:
    return "\n".join(f"{t.language}|{t.text}" for t in word.translations.all())


def _join_examples(word) -> str:
    return "\n".join(
        "|".join([e.text_arabic, e.text_cyrillic, e.translation, e.source])
        for e in word.examples.all()
    )


def export_words_xlsx(words) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "words"
    ws.append(COLUMNS)
    for w in words:
        ws.append([
            w.headword_arabic,
            w.transliteration_latin,
            w.transliteration_cyrillic,
            w.pronunciation,
            w.part_of_speech,
            w.root.morpheme if w.root else "",
            w.root.transliteration if w.root else "",
            w.root.origin if w.root else "",
            w.definition,
            w.etymology,
            w.notes,
            w.status,
            _join_translations(w),
            _join_examples(w),
        ])
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# --- import -----------------------------------------------------------------

def _parse_translations(cell: str):
    out = []
    for line in (cell or "").splitlines():
        line = line.strip()
        if not line:
            continue
        lang, _, text = line.partition("|")
        text = text.strip()
        if not text:
            continue
        lang = lang.strip()
        if lang not in _LANG_VALUES:
            lang = Translation.Language.RUSSIAN
        out.append((lang, text))
    return out


def _parse_examples(cell: str):
    out = []
    for line in (cell or "").splitlines():
        if not line.strip():
            continue
        parts = (line.split("|") + ["", "", "", ""])[:4]
        arabic, cyrillic, translation, source = (p.strip() for p in parts)
        if arabic:
            out.append((arabic, cyrillic, translation, source))
    return out


def _get_or_create_root(morpheme, translit, origin):
    morpheme = (morpheme or "").strip()
    if not morpheme:
        return None
    origin = (origin or "").strip()
    if origin not in _ORIGIN_VALUES:
        origin = Root.Origin.TURKIC
    root, _ = Root.objects.get_or_create(
        morpheme=morpheme,
        defaults={"transliteration": (translit or "").strip(), "origin": origin},
    )
    return root


def import_words_xlsx(file, user=None) -> dict:
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return {"created": 0, "updated": 0, "errors": ["Пустой файл"]}

    # Map header name -> column index (tolerant to order / extra columns).
    index = {}
    for i, name in enumerate(header):
        if name:
            index[str(name).strip().lower()] = i

    if "headword_arabic" not in index:
        return {
            "created": 0,
            "updated": 0,
            "errors": ["Нет обязательной колонки 'headword_arabic'"],
        }

    def cell(row, name):
        i = index.get(name)
        if i is None or i >= len(row) or row[i] is None:
            return ""
        return str(row[i]).strip()

    created = updated = 0
    errors = []

    for n, row in enumerate(rows, start=2):  # header is row 1
        if row is None or not any(row):
            continue
        headword = cell(row, "headword_arabic")
        if not headword:
            continue
        try:
            pos = cell(row, "part_of_speech") or PartOfSpeech.OTHER
            if pos not in _POS_VALUES:
                pos = PartOfSpeech.OTHER
            status = cell(row, "status") or Word.Status.PUBLISHED
            if status not in _STATUS_VALUES:
                status = Word.Status.PUBLISHED
            root = _get_or_create_root(
                cell(row, "root_morpheme"),
                cell(row, "root_translit"),
                cell(row, "root_origin"),
            )

            word, was_created = Word.objects.update_or_create(
                headword_arabic=headword,
                defaults={
                    "transliteration_latin": cell(row, "transliteration_latin"),
                    "transliteration_cyrillic": cell(row, "transliteration_cyrillic"),
                    "pronunciation": cell(row, "pronunciation"),
                    "part_of_speech": pos,
                    "root": root,
                    "definition": cell(row, "definition"),
                    "etymology": cell(row, "etymology"),
                    "notes": cell(row, "notes"),
                    "status": status,
                },
            )
            if was_created and user is not None:
                word.created_by = user
                word.save(update_fields=["created_by"])

            # Replace nested collections to match the row.
            word.translations.all().delete()
            Translation.objects.bulk_create([
                Translation(word=word, language=lang, text=text, order=i)
                for i, (lang, text) in enumerate(_parse_translations(cell(row, "translations")))
            ])
            word.examples.all().delete()
            UsageExample.objects.bulk_create([
                UsageExample(
                    word=word, text_arabic=ar, text_cyrillic=cy,
                    translation=tr, source=src, order=i,
                )
                for i, (ar, cy, tr, src) in enumerate(_parse_examples(cell(row, "examples")))
            ])

            created += 1 if was_created else 0
            updated += 0 if was_created else 1
        except Exception as exc:  # noqa: BLE001 — собираем ошибки построчно
            errors.append(f"строка {n}: {exc}")

    return {"created": created, "updated": updated, "errors": errors}
